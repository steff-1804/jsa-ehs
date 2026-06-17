from flask import Flask, render_template, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
from PIL import Image
import base64
import json

app = Flask(__name__)
# Límite realista para Render Free. Las fotos se optimizan en JavaScript antes de enviarse.
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024


@app.route("/")
def index():
    return render_template("index.html")


def _b64_to_image_stream(data_url: str):
    if not data_url:
        return None
    encoded = data_url.split(",", 1)[1] if "," in data_url else data_url
    raw = base64.b64decode(encoded)
    return BytesIO(raw)


def _rpn_fill(value: int):
    if value <= 10:
        return PatternFill("solid", fgColor="C6EFCE")
    if value <= 30:
        return PatternFill("solid", fgColor="FFEB9C")
    return PatternFill("solid", fgColor="FFC7CE")


def _safe_int(value, default=1):
    try:
        return int(value)
    except Exception:
        return default


def _join_controls(peligros, field):
    values = []
    for i, p in enumerate(peligros, start=1):
        text = (p.get(field) or "").strip()
        if text:
            values.append(f"Peligro {i}: {text}")
    return "\n".join(values)


def _insert_image(ws, fotos, anchor_cell, start_row, end_row):
    if not fotos:
        return
    try:
        img_stream = _b64_to_image_stream(fotos[0])
        if img_stream is None:
            return

        pil_img = Image.open(img_stream).convert("RGB")
        original_w, original_h = pil_img.size

        # Mantiene proporción. No deforma. Tamaño razonable para Excel/Render.
        max_w, max_h = 520, 360
        scale = min(max_w / original_w, max_h / original_h, 1)
        new_w = max(1, int(original_w * scale))
        new_h = max(1, int(original_h * scale))

        if scale < 1:
            pil_img = pil_img.resize((new_w, new_h), Image.LANCZOS)

        out = BytesIO()
        pil_img.save(out, format="JPEG", quality=88, optimize=True)
        out.seek(0)

        xl_img = ExcelImage(out)
        xl_img.width = new_w
        xl_img.height = new_h
        ws.add_image(xl_img, anchor_cell)

        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 18, min(new_w / 7, 78))

        total_rows = max(1, end_row - start_row + 1)
        total_height_points = new_h * 0.75
        height_each = max(80, total_height_points / total_rows)
        for r in range(start_row, end_row + 1):
            ws.row_dimensions[r].height = height_each
    except Exception as exc:
        ws[anchor_cell] = f"Imagen no válida: {exc}"


@app.route("/exportar_excel", methods=["POST"])
def exportar_excel():
    try:
        info = json.loads(request.form.get("info_general", "{}"))
        actividades = json.loads(request.form.get("actividades", "[]"))

        if not actividades:
            return jsonify({"error": "No hay actividades para exportar."}), 400

        wb = Workbook()
        ws = wb.active
        ws.title = "Matriz JSA"

        thin = Side(border_style="thin", color="A6A6A6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        section_fill = PatternFill("solid", fgColor="D9EAF7")
        info_fill = PatternFill("solid", fgColor="F8FBFF")

        # Título
        ws.merge_cells("A1:N1")
        ws["A1"] = "MATRIZ JSA - ANÁLISIS SEGURO DE TRABAJO"
        ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        info_rows = [
            ("Área", info.get("area", "")),
            ("Miembros", info.get("miembros", "")),
            ("Fecha", info.get("fecha", "")),
            ("Descripción", info.get("descripcion", "")),
        ]
        row = 2
        for label, value in info_rows:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = section_fill
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).fill = info_fill
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=2).alignment = Alignment(vertical="center", wrap_text=True)
            ws.row_dimensions[row].height = 28
            row += 1

        headers = [
            "Actividad", "Foto", "Peligro", "Consecuencia", "SEV", "LIKL",
            "Controles existentes", "CONT", "RPN", "Controles recomendados",
            "SEV Post", "LIKL Post", "CONT Post", "RPN Post"
        ]
        header_row = 7
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[header_row].height = 42

        current_row = header_row + 1
        for act in actividades:
            peligros = act.get("peligros", []) or []
            if not peligros:
                peligros = [{"peligro": "", "consecuencia": ""}]

            start_act = current_row
            end_act = current_row + len(peligros) - 1

            existing_merged = _join_controls(peligros, "existing_controls")
            recommended_merged = _join_controls(peligros, "recommended_controls")

            for idx, p in enumerate(peligros):
                sev = _safe_int(p.get("sev"), 1)
                likl = _safe_int(p.get("likl"), 1)
                cont = _safe_int(p.get("cont"), 1)
                rpn = sev * likl * cont
                sev_post = _safe_int(p.get("sev_post"), 1)
                likl_post = _safe_int(p.get("likl_post"), 1)
                cont_post = _safe_int(p.get("cont_post"), 1)
                rpn_post = sev_post * likl_post * cont_post

                values = [
                    act.get("actividad", "") if idx == 0 else "",
                    "",
                    p.get("peligro", ""),
                    p.get("consecuencia", ""),
                    sev,
                    likl,
                    existing_merged if idx == 0 else "",
                    cont,
                    rpn,
                    recommended_merged if idx == 0 else "",
                    sev_post,
                    likl_post,
                    cont_post,
                    rpn_post,
                ]

                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal="center" if col in [1,2,5,6,8,9,11,12,13,14] else "left",
                        vertical="center",
                        wrap_text=True,
                    )
                ws.cell(current_row, 9).fill = _rpn_fill(rpn)
                ws.cell(current_row, 9).font = Font(bold=True)
                ws.cell(current_row, 14).fill = _rpn_fill(rpn_post)
                ws.cell(current_row, 14).font = Font(bold=True)
                ws.row_dimensions[current_row].height = 80
                current_row += 1

            # Combinar verticalmente: Actividad, Foto, Controles existentes y recomendados.
            for col in [1, 2, 7, 10]:
                if end_act > start_act:
                    ws.merge_cells(start_row=start_act, start_column=col, end_row=end_act, end_column=col)
                cell = ws.cell(start_act, col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            _insert_image(ws, act.get("fotos", []), f"B{start_act}", start_act, end_act)

        widths = {
            "A": 28, "B": 45, "C": 32, "D": 38, "E": 10, "F": 10,
            "G": 42, "H": 10, "I": 10, "J": 42, "K": 10, "L": 10,
            "M": 10, "N": 10,
        }
        for col, width in widths.items():
            if col != "B":
                ws.column_dimensions[col].width = width
        ws.freeze_panes = "A8"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="Matriz_JSA.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


if __name__ == "__main__":
    app.run(debug=False)
